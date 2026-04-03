"""
excel_reporter.py - Модуль для формирования Excel отчетов по экспериментам
Поддерживает новую структуру JSON файлов:
- Основной JSON с результатами и конфигурацией
- Отдельный JSON с краткой сводкой
- Отдельный JSON с полной конфигурацией
"""
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell




# Импортируем конфигурацию
from config import logging_config    
from config import model_config

# Настройка логирования
logger = logging.getLogger(__name__)

#logging.basicConfig(level=logging.INFO)
#logger = logging.getLogger(__name__)


class ExcelReporter:
    """
    Класс для формирования Excel отчетов по экспериментам
    Поддерживает новую структуру файлов
    """
    
    def __init__(self, experiments_dir: Path = Path("experiments")):
        """
        Инициализация репортера
        
        Args:
            experiments_dir: Директория с экспериментами
        """
        self.experiments_dir = Path(experiments_dir)
        self.results_dir = self.experiments_dir / "results"
        self.output_file = self.experiments_dir / "experiments_summary.xlsx"

    def debug_json_structure(self, json_file: Path):
        """
        Отладочный метод для просмотра структуры JSON файла
        """
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            print(f"\n=== Структура {json_file.name} ===")
            print(f"Ключи верхнего уровня: {list(data.keys())}")
            
            if 'retrieval_metrics' in data:
                rm = data['retrieval_metrics']
                print(f"\nretrieval_metrics keys: {list(rm.keys())}")
                
                for key in ['precision_at_k', 'recall_at_k', 'ndcg_at_k']:
                    if key in rm:
                        val = rm[key]
                        print(f"{key}: {val} (type: {type(val)}, length: {len(val) if isinstance(val, list) else 'N/A'})")
            
            # Добавляем вывод advanced_metrics для отладки
            if 'advanced_metrics' in data:
                adv = data['advanced_metrics']
                print(f"\nadvanced_metrics keys: {list(adv.keys())}")
                print(f"faithfulness: {adv.get('faithfulness', 0)}")
                print(f"answer_relevancy: {adv.get('answer_relevancy', 0)}")
                print(f"context_relevancy: {adv.get('context_relevancy', 0)}")
            
            print("=" * 50)
        except Exception as e:
            print(f"Ошибка при анализе {json_file.name}: {e}")
                    
    def collect_all_results(self) -> List[Dict[str, Any]]:
        """
        Сбор всех результатов экспериментов из JSON файлов
        """
        results = []

        if not self.results_dir.exists():
            logger.error(f"Директория {self.results_dir} не найдена")
            return results

        # Находим все основные JSON файлы с результатами экспериментов
        json_files = list(self.results_dir.glob("my_experiment_*.json"))
        json_files.extend(self.results_dir.glob("*_experiment_*.json"))

        # Фильтруем только основные файлы (не summary и не config)
        main_json_files = []
        for json_file in json_files:
            if '_summary' not in json_file.stem and '_config' not in json_file.stem:
                main_json_files.append(json_file)

        logger.info(f"Найдено {len(main_json_files)} основных файлов с результатами")

        # Отладочный вывод для первого файла
        if main_json_files:
            self.debug_json_structure(main_json_files[0])

        for json_file in main_json_files:
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Пытаемся загрузить дополнительную конфигурацию если есть
                config_data = self._load_configuration(json_file)

                # Извлекаем основные данные
                experiment_data = self._extract_experiment_data(data, config_data, json_file.stem)
                results.append(experiment_data)
                logger.info(f"Обработан: {json_file.name}")

            except Exception as e:
                logger.error(f"Ошибка при обработке {json_file.name}: {e}")

        return results

    def _load_configuration(self, json_file: Path) -> Optional[Dict[str, Any]]:
        """
        Загрузка конфигурации из отдельного файла если существует
        
        Args:
            json_file: Путь к основному JSON файлу
            
        Returns:
            Dict: Конфигурация или None
        """
        # Пробуем загрузить конфигурацию из отдельного файла
        config_file = json_file.parent / f"{json_file.stem}_config.json"
        
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    logger.info(f"  Загружена конфигурация из: {config_file.name}")
                    return json.load(f)
            except Exception as e:
                logger.warning(f"  Не удалось загрузить конфигурацию: {e}")
        
        return None

    def _extract_experiment_data(self, 
                                 data: Dict[str, Any], 
                                 config_data: Optional[Dict[str, Any]],
                                 experiment_id: str) -> Dict[str, Any]:
        """
        Извлечение данных из JSON эксперимента
        """
        # Базовая информация
        experiment_info = {
            'experiment_id': experiment_id,
            'timestamp': data.get('timestamp', ''),
            'status': data.get('status', 'unknown'),
        }

        # Конфигурация эксперимента (из основного файла)
        config = data.get('config', {})

        # Если есть отдельная конфигурация, используем её для расширенных данных
        system_config = config_data.get('system_configuration', {}) if config_data else {}

        experiment_info.update({
            'experiment_name': config.get('experiment_name', ''),
            'retrieval_type': config.get('retrieval_type', ''),
            'k_retrieve': config.get('k_retrieve', 0),
            'num_test_samples': config.get('num_test_samples', 0),
            'num_errors': config.get('num_errors', 0),
        })

        # Извлекаем информацию о моделях
        llm_type = config.get('llm_model', '')
        if not llm_type:
            llm_type = system_config.get('model_config', {}).get('llm_model_name', '')
        if not llm_type:
            llm_type = config.get('llm_model_name', 'unknown')

        embedding_type = config.get('embedding_model', '')
        if not embedding_type:
            embedding_type = system_config.get('model_config', {}).get('embedding_model_name', '')
        if not embedding_type:
            embedding_type = config.get('embedding_model_name', 'unknown')

        chunk_size = config.get('chunk_size', 0)
        if not chunk_size:
            chunk_size = system_config.get('data_config', {}).get('chunk_size', 0)

        chunk_overlap = config.get('chunk_overlap', 0)
        if not chunk_overlap:
            chunk_overlap = system_config.get('data_config', {}).get('chunk_overlap', 0)

        temperature = config.get('temperature', 0)
        if not temperature:
            temperature = system_config.get('model_config', {}).get('temperature', 0)

        device = config.get('device', '')
        if not device:
            device = system_config.get('model_config', {}).get('device', '')

        experiment_info.update({
            'llm_type': llm_type,
            'embedding_type': embedding_type,
            'chunk_size': chunk_size,
            'chunk_overlap': chunk_overlap,
            'temperature': temperature,
            'device': device,
        })

        # МЕТРИКИ ПОИСКА
        retrieval_metrics = data.get('retrieval_metrics', {})

        # Базовые метрики
        experiment_info['map_score'] = float(retrieval_metrics.get('map', 0.0))
        experiment_info['mrr_score'] = float(retrieval_metrics.get('mrr', 0.0))

        # Функция для извлечения значений из метрик
        def get_metric_value(metric_data, key):
            """Извлекает значение из метрик"""
            if isinstance(metric_data, dict):
                return float(metric_data.get(key, 0.0))
            elif isinstance(metric_data, list) and isinstance(key, int) and len(metric_data) > key:
                return float(metric_data[key])
            return 0.0

        # Маппинг k значений к ключам в словаре
        k_index_map = {1: '1', 3: '3', 5: '5', 10: '10'}

        # Извлекаем Precision@k, Recall@k, NDCG@k
        for k, key in k_index_map.items():
            experiment_info[f'precision_at_{k}'] = get_metric_value(
                retrieval_metrics.get('precision_at_k', {}), key
            )
            experiment_info[f'recall_at_{k}'] = get_metric_value(
                retrieval_metrics.get('recall_at_k', {}), key
            )
            experiment_info[f'ndcg_at_{k}'] = get_metric_value(
                retrieval_metrics.get('ndcg_at_k', {}), key
            )

        # Для отладки - выводим полученные значения
        logger.info(f"Извлеченные метрики для {experiment_id}:")
        logger.info(f"  MAP: {experiment_info['map_score']}, MRR: {experiment_info['mrr_score']}")
        logger.info(f"  P@1: {experiment_info['precision_at_1']}, P@3: {experiment_info['precision_at_3']}")
        logger.info(f"  R@1: {experiment_info['recall_at_1']}, R@3: {experiment_info['recall_at_3']}")

        # Метрики генерации
        generation_metrics = data.get('generation_metrics', {})
        experiment_info.update({
            'rouge1': generation_metrics.get('rouge1', 0.0),
            'rouge2': generation_metrics.get('rouge2', 0.0),
            'rougeL': generation_metrics.get('rougeL', 0.0),
            'bleu': generation_metrics.get('bleu', 0.0),
            'bert_score_f1': generation_metrics.get('bert_score_f1', 0.0),
        })

        # ПРОДВИНУТЫЕ МЕТРИКИ RAGAS
        advanced_metrics = data.get('advanced_metrics', {})
        if advanced_metrics:
            experiment_info.update({
                'faithfulness': advanced_metrics.get('faithfulness', 0.0),
                'answer_relevancy': advanced_metrics.get('answer_relevancy', 0.0),
                'context_relevancy': advanced_metrics.get('context_relevancy', 0.0),
            })

        # Метрики производительности
        experiment_info.update({
            'execution_time_sec': data.get('execution_time', 0.0),
            'avg_retrieval_time': config.get('avg_retrieval_time', 0.0),
            'avg_generation_time': config.get('avg_generation_time', 0.0),
        })

        # Дополнительная информация из системной конфигурации
        if system_config:
            vector_config = system_config.get('vectorstore_config', {})
            experiment_info.update({
                'vector_store_type': vector_config.get('vector_store_type', ''),
                'faiss_index_type': vector_config.get('faiss_index_type', ''),
            })

            gigachat_config = system_config.get('gigachat_config', {})
            if gigachat_config.get('enabled'):
                experiment_info['gigachat_model'] = gigachat_config.get('model', '')

        # 🔥 ДОБАВИТЬ ИЗВЛЕЧЕНИЕ ТОКЕНОВ
        token_stats = data.get('token_stats', {})
        if token_stats:
            experiment_info.update({
                'total_tokens': token_stats.get('total_tokens', 0),
                'total_prompt_tokens': token_stats.get('total_prompt_tokens', 0),
                'total_completion_tokens': token_stats.get('total_completion_tokens', 0),
                'num_requests': token_stats.get('num_requests', 0),
                'avg_tokens_per_request': token_stats.get('avg_tokens_per_request', 0),
                'estimated_cost_usd': token_stats.get('estimated_cost_usd', 0),
            })

        # Извлекаем статистику баланса
        balance_stats = data.get('balance_stats', {})
        if balance_stats:
            total_delta = balance_stats.get('total_delta', {})
            experiment_info.update({
                'initial_balance': total_delta.get('balance', 0),
                'final_balance': balance_stats.get('last_balance', {}).get('balance', 0),
                'balance_spent': total_delta.get('balance', 0),
                'balance_checks': balance_stats.get('num_checks', 0)
            })

        return experiment_info
    
    def create_summary_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Создание сводного DataFrame из результатов
        
        Args:
            results: Список результатов экспериментов
            
        Returns:
            pd.DataFrame: Сводная таблица
        """
        if not results:
            return pd.DataFrame()
        
        df = pd.DataFrame(results)
        
        # Сортируем по времени (новые сверху)
        if 'timestamp' in df.columns:
            df = df.sort_values('timestamp', ascending=False)
        
        return df
    
    def create_detailed_excel(self, df: pd.DataFrame, output_path: Path):
        """
        Создание детального Excel файла с группировкой столбцов
        
        Args:
            df: DataFrame с результатами
            output_path: Путь для сохранения
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Создаем лист с группировкой
            self._write_grouped_summary_sheet(df, writer)
            
            # Дополнительные листы
            self._write_retrieval_metrics_sheet(df, writer)
            self._write_generation_metrics_sheet(df, writer)
            self._write_performance_metrics_sheet(df, writer)
            self._write_configurations_sheet(df, writer)
            self._write_best_experiments_sheet(df, writer)
            self._write_advanced_metrics_sheet(df, writer)
            self._write_token_usage_sheet(df, writer)
            self._write_balance_sheet(df, writer)

        # Применяем стили
        self._apply_styling(output_path)
        
        logger.info(f"✅ Excel отчет создан: {output_path}")


    # Добавить лист с балансом
    def _write_balance_sheet(self, df: pd.DataFrame, writer):
        """Запись информации о балансе"""
        balance_cols = ['experiment_name', 'initial_balance', 'final_balance', 
                        'balance_spent', 'balance_checks']

        available_cols = [col for col in balance_cols if col in df.columns]

        if available_cols:
            balance_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'initial_balance': 'Начальный баланс',
                'final_balance': 'Конечный баланс',
                'balance_spent': 'Потрачено',
                'balance_checks': 'Проверок баланса'
            }
            balance_df = balance_df.rename(columns=column_names)
            balance_df.to_excel(writer, sheet_name='Баланс GigaChat', index=False)
            logger.info("  ✅ Лист 'Баланс GigaChat' создан")

    def _write_token_usage_sheet(self, df: pd.DataFrame, writer):
        """
        Запись информации об использовании токенов
        """
        token_cols = ['experiment_name', 'total_tokens', 'total_prompt_tokens', 
                      'total_completion_tokens', 'num_requests', 
                      'avg_tokens_per_request', 'estimated_cost_usd']

        available_cols = [col for col in token_cols if col in df.columns]

        if available_cols and len(available_cols) > 1:
            token_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'total_tokens': 'Всего токенов',
                'total_prompt_tokens': 'Prompt токены',
                'total_completion_tokens': 'Completion токены',
                'num_requests': 'Кол-во запросов',
                'avg_tokens_per_request': 'Ср. токенов на запрос',
                'estimated_cost_usd': 'Оценочная стоимость ($)'
            }
            token_df = token_df.rename(columns=column_names)
            token_df.to_excel(writer, sheet_name='Использование токенов', index=False)
            logger.info("  ✅ Лист 'Использование токенов' создан")

    def _write_advanced_metrics_sheet(self, df: pd.DataFrame, writer):
        """
        Запись продвинутых метрик RAGAS
        """
        advanced_cols = ['experiment_name', 'faithfulness', 'answer_relevancy', 'context_relevancy']
        available_cols = [col for col in advanced_cols if col in df.columns]
        
        if available_cols and len(available_cols) > 1:
            adv_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'faithfulness': 'Faithfulness',
                'answer_relevancy': 'Answer Relevancy',
                'context_relevancy': 'Context Relevancy'
            }
            adv_df = adv_df.rename(columns=column_names)
            adv_df.to_excel(writer, sheet_name='Продвинутые метрики', index=False)
            logger.info("  ✅ Лист 'Продвинутые метрики' создан")
                
    def _write_grouped_summary_sheet(self, df: pd.DataFrame, writer):
        """
        Запись общей сводки с группировкой столбцов и двухстрочным заголовком
        
        Args:
            df: DataFrame с результатами
            writer: Excel writer
        """
        if df.empty:
            return
        
        # Определяем структуру блоков с учетом RAGAS метрик
        blocks = {
            'Эксперимент': {
                'start_col': 0,
                'columns': [
                    ('experiment_name', 'Название эксперимента'),
                    ('timestamp', 'Время проведения'),
                    ('status', 'Статус')
                ]
            },
            'Метрики поиска': {
                'start_col': 3,
                'columns': [
                    ('map_score', 'MAP'),
                    ('mrr_score', 'MRR'),
                    ('precision_at_1', 'P@1'),
                    ('precision_at_3', 'P@3'),
                    ('precision_at_5', 'P@5'),
                    ('precision_at_10', 'P@10'),
                    ('recall_at_1', 'R@1'),
                    ('recall_at_3', 'R@3'),
                    ('recall_at_5', 'R@5'),
                    ('recall_at_10', 'R@10'),
                    ('ndcg_at_1', 'NDCG@1'),
                    ('ndcg_at_3', 'NDCG@3'),
                    ('ndcg_at_5', 'NDCG@5'),
                    ('ndcg_at_10', 'NDCG@10')
                ]
            },
            'Метрики генерации': {
                'start_col': 17,
                'columns': [
                    ('rouge1', 'ROUGE-1'),
                    ('rouge2', 'ROUGE-2'),
                    ('rougeL', 'ROUGE-L'),
                    ('bleu', 'BLEU'),
                    ('bert_score_f1', 'BERTScore F1')
                ]
            },
            'RAGAS метрики': {
                'start_col': 22,  # 17 + 5 (количество колонок в метриках генерации)
                'columns': [
                    ('faithfulness', 'Faithfulness'),
                    ('answer_relevancy', 'Answer Relevancy'),
                    ('context_relevancy', 'Context Relevancy')
                ]
            },
            'Производительность': {
                'start_col': 25,  # 22 + 3 (количество колонок в RAGAS метриках)
                'columns': [
                    ('execution_time_sec', 'Общее время (сек)'),
                    ('avg_retrieval_time', 'Ср. время поиска (сек)'),
                    ('avg_generation_time', 'Ср. время генерации (сек)'),
                    ('num_test_samples', 'Кол-во тестов'),
                    ('num_errors', 'Кол-во ошибок')
                ]
            },
            'Конфигурации': {
                'start_col': 30,  # 25 + 5 (количество колонок в производительности)
                'columns': [
                    ('retrieval_type', 'Тип поиска'),
                    ('llm_type', 'Тип LLM'),
                    ('embedding_type', 'Эмбеддинги'),
                    ('chunk_size', 'Chunk size'),
                    ('chunk_overlap', 'Chunk overlap'),
                    ('temperature', 'Temperature'),
                    ('k_retrieve', 'K (кол-во документов)')
                ]
            },
            'Использование токенов': {
                'start_col': 37,  # Рассчитайте правильную позицию
                'columns': [
                    ('total_tokens', 'Всего токенов'),
                    ('total_prompt_tokens', 'Prompt токены'),
                    ('total_completion_tokens', 'Completion токены'),
                    ('num_requests', 'Кол-во запросов'),
                    ('avg_tokens_per_request', 'Ср. токенов/запрос'),
                    ('estimated_cost_usd', 'Стоимость ($)')
                ]
            }            
        }
        
        # Подготавливаем данные для вывода
        data_rows = []
        for _, row in df.iterrows():
            data_row = []
            for block_name, block_info in blocks.items():
                for col_key, _ in block_info['columns']:
                    value = row.get(col_key, '')
                    # Форматирование чисел
                    if isinstance(value, float):
                        value = round(value, 4)
                    elif isinstance(value, (int, float)) and value == 0:
                        value = 0
                    data_row.append(value)
            data_rows.append(data_row)
        
        if not data_rows:
            logger.warning("Нет данных для записи в Excel")
            return
        
        # Создаем лист
        worksheet = writer.book.create_sheet("Общая сводка")
        
        # Определяем границы блоков
        block_boundaries = {}
        current_col = 0
        for block_name, block_info in blocks.items():
            block_start = current_col
            block_end = current_col + len(block_info['columns']) - 1
            block_boundaries[block_name] = (block_start, block_end)
            current_col = block_end + 1
        
        # Заполняем первую строку (названия блоков)
        for block_name, (start_col, end_col) in block_boundaries.items():
            cell = worksheet.cell(row=1, column=start_col + 1)
            cell.value = block_name
            # Объединяем ячейки для блока
            if start_col != end_col:
                worksheet.merge_cells(
                    start_row=1, 
                    start_column=start_col + 1,
                    end_row=1, 
                    end_column=end_col + 1
                )
        
        # Заполняем вторую строку (названия колонок)
        col_idx = 0
        for block_name, block_info in blocks.items():
            for _, col_name in block_info['columns']:
                cell = worksheet.cell(row=2, column=col_idx + 1)
                cell.value = col_name
                col_idx += 1
        
        # Заполняем данные
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data):
                cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                cell.value = value
        
        # Настройка ширины колонок
        for col_idx in range(len(data_rows[0])):
            column_letter = get_column_letter(col_idx + 1)
            worksheet.column_dimensions[column_letter].width = 18
        
        # Заморозка панели (первые две строки)
        worksheet.freeze_panes = 'A3'
        
        logger.info("  ✅ Лист 'Общая сводка' создан")
    
    def _write_retrieval_metrics_sheet(self, df: pd.DataFrame, writer):
        """Запись метрик поиска"""
        retrieval_cols = [
            'experiment_name', 'retrieval_type', 'k_retrieve',
            'map_score', 'mrr_score',
            'precision_at_1', 'precision_at_3', 'precision_at_5', 'precision_at_10',
            'recall_at_1', 'recall_at_3', 'recall_at_5', 'recall_at_10',
            'ndcg_at_1', 'ndcg_at_3', 'ndcg_at_5', 'ndcg_at_10'
        ]
        
        available_cols = [col for col in retrieval_cols if col in df.columns]
        if available_cols:
            retrieval_df = df[available_cols].copy()
            
            column_names = {
                'experiment_name': 'Эксперимент',
                'retrieval_type': 'Тип поиска',
                'k_retrieve': 'K',
                'map_score': 'MAP',
                'mrr_score': 'MRR',
                'precision_at_1': 'P@1',
                'precision_at_3': 'P@3',
                'precision_at_5': 'P@5',
                'precision_at_10': 'P@10',
                'recall_at_1': 'R@1',
                'recall_at_3': 'R@3',
                'recall_at_5': 'R@5',
                'recall_at_10': 'R@10',
                'ndcg_at_1': 'NDCG@1',
                'ndcg_at_3': 'NDCG@3',
                'ndcg_at_5': 'NDCG@5',
                'ndcg_at_10': 'NDCG@10'
            }
            
            retrieval_df = retrieval_df.rename(columns=column_names)
            retrieval_df.to_excel(writer, sheet_name='Метрики поиска', index=False)
            logger.info("  ✅ Лист 'Метрики поиска' создан")
    
    def _write_generation_metrics_sheet(self, df: pd.DataFrame, writer):
        """Запись метрик генерации"""
        gen_cols = ['experiment_name', 'rouge1', 'rouge2', 'rougeL', 'bleu', 'bert_score_f1']
        available_cols = [col for col in gen_cols if col in df.columns]
        
        if available_cols:
            gen_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'rouge1': 'ROUGE-1',
                'rouge2': 'ROUGE-2',
                'rougeL': 'ROUGE-L',
                'bleu': 'BLEU',
                'bert_score_f1': 'BERTScore F1'
            }
            gen_df = gen_df.rename(columns=column_names)
            gen_df.to_excel(writer, sheet_name='Метрики генерации', index=False)
            logger.info("  ✅ Лист 'Метрики генерации' создан")
    
    def _write_performance_metrics_sheet(self, df: pd.DataFrame, writer):
        """Запись метрик производительности"""
        perf_cols = ['experiment_name', 'execution_time_sec', 'avg_retrieval_time', 
                    'avg_generation_time', 'num_test_samples', 'num_errors']
        available_cols = [col for col in perf_cols if col in df.columns]
        
        if available_cols:
            perf_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'execution_time_sec': 'Общее время (сек)',
                'avg_retrieval_time': 'Ср. время поиска (сек)',
                'avg_generation_time': 'Ср. время генерации (сек)',
                'num_test_samples': 'Кол-во тестов',
                'num_errors': 'Кол-во ошибок'
            }
            perf_df = perf_df.rename(columns=column_names)
            perf_df.to_excel(writer, sheet_name='Производительность', index=False)
            logger.info("  ✅ Лист 'Производительность' создан")
    
    def _write_configurations_sheet(self, df: pd.DataFrame, writer):
        """Запись конфигураций экспериментов"""
        config_cols = ['experiment_name', 'retrieval_type', 'llm_type', 'embedding_type', 
                      'chunk_size', 'chunk_overlap', 'temperature', 'device',
                      'k_retrieve', 'num_test_samples', 'timestamp']
        available_cols = [col for col in config_cols if col in df.columns]
        
        if available_cols:
            config_df = df[available_cols].copy()
            column_names = {
                'experiment_name': 'Эксперимент',
                'retrieval_type': 'Тип поиска',
                'llm_type': 'Тип LLM',
                'embedding_type': 'Эмбеддинги',
                'chunk_size': 'Chunk size',
                'chunk_overlap': 'Chunk overlap',
                'temperature': 'Temperature',
                'device': 'Device',
                'k_retrieve': 'K',
                'num_test_samples': 'Тестов',
                'timestamp': 'Время'
            }
            config_df = config_df.rename(columns=column_names)
            config_df.to_excel(writer, sheet_name='Конфигурации', index=False)
            logger.info("  ✅ Лист 'Конфигурации' создан")
    
    def _write_best_experiments_sheet(self, df: pd.DataFrame, writer):
        """Запись лучших экспериментов по разным метрикам"""
        if df.empty:
            return
        
        best_results = []
        
        # Лучший по MAP
        if 'map_score' in df.columns and not df['map_score'].isna().all() and len(df) > 0:
            best_idx = df['map_score'].idxmax()
            best_map = df.loc[best_idx]
            best_results.append({
                'Метрика': 'MAP',
                'Эксперимент': best_map.get('experiment_name', ''),
                'Значение': round(best_map.get('map_score', 0), 4),
                'Тип поиска': best_map.get('retrieval_type', ''),
                'LLM': best_map.get('llm_type', ''),
                'Эмбеддинги': best_map.get('embedding_type', '')
            })
        
        # Лучший по MRR
        if 'mrr_score' in df.columns and not df['mrr_score'].isna().all() and len(df) > 0:
            best_idx = df['mrr_score'].idxmax()
            best_mrr = df.loc[best_idx]
            best_results.append({
                'Метрика': 'MRR',
                'Эксперимент': best_mrr.get('experiment_name', ''),
                'Значение': round(best_mrr.get('mrr_score', 0), 4),
                'Тип поиска': best_mrr.get('retrieval_type', ''),
                'LLM': best_mrr.get('llm_type', ''),
                'Эмбеддинги': best_mrr.get('embedding_type', '')
            })
        
        # Лучший по ROUGE-1
        if 'rouge1' in df.columns and not df['rouge1'].isna().all() and len(df) > 0:
            best_idx = df['rouge1'].idxmax()
            best_rouge = df.loc[best_idx]
            best_results.append({
                'Метрика': 'ROUGE-1',
                'Эксперимент': best_rouge.get('experiment_name', ''),
                'Значение': round(best_rouge.get('rouge1', 0), 4),
                'Тип поиска': best_rouge.get('retrieval_type', ''),
                'LLM': best_rouge.get('llm_type', ''),
                'Эмбеддинги': best_rouge.get('embedding_type', '')
            })
        
        # Лучший по BLEU
        if 'bleu' in df.columns and not df['bleu'].isna().all() and len(df) > 0:
            best_idx = df['bleu'].idxmax()
            best_bleu = df.loc[best_idx]
            best_results.append({
                'Метрика': 'BLEU',
                'Эксперимент': best_bleu.get('experiment_name', ''),
                'Значение': round(best_bleu.get('bleu', 0), 4),
                'Тип поиска': best_bleu.get('retrieval_type', ''),
                'LLM': best_bleu.get('llm_type', ''),
                'Эмбеддинги': best_bleu.get('embedding_type', '')
            })
        
        # Лучший по Faithfulness (RAGAS)
        if 'faithfulness' in df.columns and not df['faithfulness'].isna().all() and len(df) > 0:
            best_idx = df['faithfulness'].idxmax()
            best_faith = df.loc[best_idx]
            best_results.append({
                'Метрика': 'Faithfulness',
                'Эксперимент': best_faith.get('experiment_name', ''),
                'Значение': round(best_faith.get('faithfulness', 0), 4),
                'Тип поиска': best_faith.get('retrieval_type', ''),
                'LLM': best_faith.get('llm_type', ''),
                'Эмбеддинги': best_faith.get('embedding_type', '')
            })
        
        # Лучший по Answer Relevancy (RAGAS)
        if 'answer_relevancy' in df.columns and not df['answer_relevancy'].isna().all() and len(df) > 0:
            best_idx = df['answer_relevancy'].idxmax()
            best_answer = df.loc[best_idx]
            best_results.append({
                'Метрика': 'Answer Relevancy',
                'Эксперимент': best_answer.get('experiment_name', ''),
                'Значение': round(best_answer.get('answer_relevancy', 0), 4),
                'Тип поиска': best_answer.get('retrieval_type', ''),
                'LLM': best_answer.get('llm_type', ''),
                'Эмбеддинги': best_answer.get('embedding_type', '')
            })
        
        # Лучший по Context Relevancy (RAGAS)
        if 'context_relevancy' in df.columns and not df['context_relevancy'].isna().all() and len(df) > 0:
            best_idx = df['context_relevancy'].idxmax()
            best_context = df.loc[best_idx]
            best_results.append({
                'Метрика': 'Context Relevancy',
                'Эксперимент': best_context.get('experiment_name', ''),
                'Значение': round(best_context.get('context_relevancy', 0), 4),
                'Тип поиска': best_context.get('retrieval_type', ''),
                'LLM': best_context.get('llm_type', ''),
                'Эмбеддинги': best_context.get('embedding_type', '')
            })
        
        # Лучший по времени выполнения
        if 'execution_time_sec' in df.columns and not df['execution_time_sec'].isna().all() and len(df) > 0:
            best_idx = df['execution_time_sec'].idxmin()
            best_time = df.loc[best_idx]
            best_results.append({
                'Метрика': 'Мин. время выполнения',
                'Эксперимент': best_time.get('experiment_name', ''),
                'Значение': round(best_time.get('execution_time_sec', 0), 2),
                'Тип поиска': best_time.get('retrieval_type', ''),
                'LLM': best_time.get('llm_type', ''),
                'Эмбеддинги': best_time.get('embedding_type', '')
            })
        
        if best_results:
            best_df = pd.DataFrame(best_results)
            best_df.to_excel(writer, sheet_name='Лучшие эксперименты', index=False)
            logger.info("  ✅ Лист 'Лучшие эксперименты' создан")
    
    def _apply_styling(self, file_path: Path):
        """
        Применение стилей к Excel файлу
        
        Args:
            file_path: Путь к Excel файлу
        """
        try:
            wb = load_workbook(file_path)
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_fill_light = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Границы
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Пропускаем применение стилей к объединенным ячейкам
                merged_cells = []
                for merged_range in ws.merged_cells.ranges:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            merged_cells.append((row, col))
                
                # Применяем стили к заголовкам
                for row in [1, 2]:
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            # Пропускаем объединенные ячейки
                            if (row, col) not in merged_cells:
                                cell.font = header_font
                                if row == 1:
                                    cell.fill = header_fill
                                else:
                                    cell.fill = header_fill_light
                                cell.alignment = header_alignment
                                cell.border = thin_border

                # Автоматическая ширина колонок (с пропуском объединенных ячеек)
                for column in ws.columns:
                    max_length = 0
                    try:
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                # Пропускаем объединенные ячейки
                                if isinstance(cell, MergedCell):
                                    continue
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    except Exception as e:
                        logger.debug(f"Не удалось установить ширину для колонки: {e}")                
                
                # Выравнивание для всех ячеек (кроме объединенных)
                for row in ws.iter_rows(min_row=3):
                    for cell in row:
                        if cell.value and (cell.row, cell.column) not in merged_cells:
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = thin_border
            
            wb.save(file_path)
            logger.info("✅ Стили применены к Excel файлу")
            
        except Exception as e:
            logger.warning(f"Не удалось применить стили: {e}")
    
    def generate_report(self) -> Path:
        """
        Генерация полного Excel отчета
        
        Returns:
            Path: Путь к созданному файлу
        """
        logger.info("📊 Начинаем генерацию Excel отчета...")
        
        # Сбор результатов
        results = self.collect_all_results()
        
        if not results:
            logger.warning("Нет результатов для генерации отчета")
            return None
        
        # Создание DataFrame
        df = self.create_summary_dataframe(results)
        
        # Создание Excel
        self.create_detailed_excel(df, self.output_file)
        
        # Вывод статистики
        logger.info(f"\n📈 Статистика:")
        logger.info(f"  - Всего экспериментов: {len(results)}")
        logger.info(f"  - Успешных: {sum(1 for r in results if r.get('status') == 'completed')}")
        logger.info(f"  - С ошибками: {sum(1 for r in results if r.get('status') == 'failed')}")
        
        return self.output_file


def generate_experiments_summary():
    """
    Основная функция для генерации Excel отчета
    """
    reporter = ExcelReporter()
    output_file = reporter.generate_report()
    
    if output_file and output_file.exists():
        print(f"\n✅ Отчет успешно создан: {output_file}")
        print(f"   Размер файла: {output_file.stat().st_size / 1024:.2f} KB")
    else:
        print("\n❌ Не удалось создать отчет")


if __name__ == "__main__":
    generate_experiments_summary()